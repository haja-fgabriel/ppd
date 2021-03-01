import os
import logging
import subprocess
import sys
from socket import *
import openpyxl

DEFAULT_IP = '127.0.0.1'
DEFAULT_PORT = 40000
PROJECT_NAME = "PolynomeAdder"
EXEC_DIRECTORY = f"{PROJECT_NAME}\\Debug\\"

ARGS_STRINGS = [
    "<excel_file>",
    "<num_polynomes>",
    "<max_degree>",
    "<num_monomes>",
    "<number_runs>"
]
NUM_ARGS = len(ARGS_STRINGS)

def log_to_excel(filename, time, num_threads, num_reading_threads, num_polynomes, max_degree, num_monomes):
    workbook = None
    if os.path.exists(filename):
        workbook = openpyxl.load_workbook(filename)
    else:
	    workbook = openpyxl.Workbook()
	    workbook.active.append(('Run no.', 'Language', 'Execution time', 'No. of threads', 'No. of reading threads', 'No. polynomes', 'Max degree', 'No. monomes'))
    sheet1 = workbook.active
    
    last_run_no = sheet1.max_row
    sheet1.append((last_run_no, 'C++', time, num_threads, num_reading_threads, num_polynomes, max_degree, num_monomes))
   
    workbook.save(filename)

if __name__ == '__main__':
    if len(sys.argv) < len(ARGS_STRINGS) + 1:
        exit(f"Usage: {sys.argv[0]} " + ' '.join(ARGS_STRINGS) + " [<num_threads> <num_reading_threads>] ")

    # special socket for getting logs from client
    srv = socket(AF_INET, SOCK_STREAM)
    srv.bind((DEFAULT_IP, DEFAULT_PORT))
    srv.listen(10)
    #print(sys.argv[6:])
	
    gen_args = [
        EXEC_DIRECTORY + "FileGenerator.exe", 
        "", 
        sys.argv[3], 
        sys.argv[4]
    ]
    for i in range(1, int(sys.argv[2]) + 1):
        gen_args[1] = EXEC_DIRECTORY + f"Polynome{i:03}.txt"
        print(gen_args)
        gen_proc = subprocess.Popen(gen_args)
        (out, err) = gen_proc.communicate()
	
    
    args = ([
		EXEC_DIRECTORY + ("Sequential" if len(sys.argv) == NUM_ARGS+1 else "") + PROJECT_NAME + ".exe", 
		sys.argv[2],
		EXEC_DIRECTORY + "result.txt",
	]  ) +  ([sys.argv[6], sys.argv[7]] if len(sys.argv) > (NUM_ARGS + 1) else [])  + [
		"--log-to-socket"
	]
	
    print(args)
    avg_time = 0.0
    for i in range(int(sys.argv[5])):
        proc = subprocess.Popen(args)
        cl, new_addr = srv.accept()
        
		# we should receive a string formatted as "time 0.001"
        receipt = cl.recv(50)
        words = receipt.decode('utf-8').strip().split()
        cl.close()
		
        # remove comment if you want to manipulate program's output
        # don't move this function before recv, otherwise 
        # a deadlock will occur, as the child process is waiting for recv in this process,
        # and this process is waiting for the child process to terminate
        (out, err) = proc.communicate()
        avg_time += float(words[1])

    avg_time /= int(sys.argv[5])

    log_to_excel(sys.argv[1], avg_time, '1' if len(sys.argv) == NUM_ARGS+1 else sys.argv[6], '1' if len(sys.argv) == NUM_ARGS+1 else sys.argv[7], sys.argv[2], sys.argv[3], sys.argv[4])