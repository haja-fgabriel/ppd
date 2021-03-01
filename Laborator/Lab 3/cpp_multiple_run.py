import os
import logging
import subprocess
import sys
from socket import *
import openpyxl

DEFAULT_IP = '127.0.0.1'
DEFAULT_PORT = 40000

def log_to_excel(filename, time, num_processes, type, N, M):
    workbook = None
    if os.path.exists(filename):
        workbook = openpyxl.load_workbook(filename)
    else:
	    workbook = openpyxl.Workbook()
	    workbook.active.append(('Run no.', 'Execution time', 'No. of processes', 'Transmission type', 'N', 'M'))
    sheet1 = workbook.active
    
    last_run_no = sheet1.max_row
    sheet1.append((last_run_no, time, num_processes, type, N, M))
   
    workbook.save(filename)

if __name__ == '__main__':
    if len(sys.argv) < 6:
        exit(f"Usage: {sys.argv[0]} <excel_file> <N1> <M2> <number_runs> sequential|scatter|isend [<num_processes>] ")

    # special socket for getting logs from client
    srv = socket(AF_INET, SOCK_STREAM)
    srv.bind((DEFAULT_IP, DEFAULT_PORT))
    srv.listen(10)
    #print(sys.argv[6:])
	
    gen_args = [
		"AdditionApp\\Debug\\FileGenerator.exe", 
		"AdditionApp\\Debug\\Number1.txt", 
		sys.argv[2], 
	] 
    gen_proc = subprocess.Popen(gen_args)
    (out, err) = gen_proc.communicate()
	
    gen_args[1] = "AdditionApp\\Debug\\Number2.txt"
    gen_args[2] = sys.argv[3]
    gen_proc = subprocess.Popen(gen_args)
	 
    (out, err) = gen_proc.communicate()
    args = ([
        "mpiexec",
        "-n",
        sys.argv[6],
        ("AdditionApp\\Debug\\MPIAddition" + ("SendRecv" if sys.argv[5] == 'isend' else ""))
    ] if (len(sys.argv) >= 7) else ["AdditionApp\\Debug\\Addition"]) + [
        "AdditionApp\\Debug\\Number1.txt",
        "AdditionApp\\Debug\\Number2.txt",
        "AdditionApp\\Debug\\Number3.txt",
		"--log-to-socket"
	]
    print(args)
    avg_time = 0.0
    for i in range(int(sys.argv[4])):
        proc = subprocess.Popen(args)
		
		# remove comment if you want to manipulate program's output
        #(out, err) = proc.communicate()
        cl, new_addr = srv.accept()
        
		# we should receive a string formatted at "time 0.001"
        receipt = cl.recv(50)
        words = receipt.decode('utf-8').strip().split()
        cl.close()
        avg_time += float(words[1])

    avg_time /= int(sys.argv[4])

    log_to_excel(sys.argv[1], avg_time, '1' if len(sys.argv) < 7 else sys.argv[6], sys.argv[5], sys.argv[2], sys.argv[3])