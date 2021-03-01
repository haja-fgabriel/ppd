import os
import logging
import subprocess
import sys
from socket import *
import openpyxl

DEFAULT_IP = '127.0.0.1'
DEFAULT_PORT = 40000

def log_to_excel(filename, time, num_threads, type, N, M, n):
    workbook = None
    if os.path.exists(filename):
        workbook = openpyxl.load_workbook(filename)
    else:
	    workbook = openpyxl.Workbook()
	    workbook.active.append(('Run no.', 'Language', 'Execution time', 'No. of threads', 'Allocation type', 'N', 'M', 'n'))
    sheet1 = workbook.active
    
    last_run_no = sheet1.max_row
    sheet1.append((last_run_no, 'Java', time, num_threads, type, N, M, n))
   
    workbook.save(filename)

if __name__ == '__main__':
    if len(sys.argv) < 6:
        exit(f"Usage: {sys.argv[0]} <excel_file> <N> <M> <n> <number_runs> [<num_threads>] ")

    # special socket for getting logs from client
    srv = socket(AF_INET, SOCK_STREAM)
    srv.bind((DEFAULT_IP, DEFAULT_PORT))
    srv.listen(10)
    #print(sys.argv[6:])
	
    gen_args = [
		"C++\PixelTransformation\Debug\FileGenerator.exe", 
		"Java\\build\\classes\\java\\main\\data.txt", 
		sys.argv[2], 
		sys.argv[3], 
		"-1000", 
		"1000"
	] 
    
    print(gen_args)
    
    gen_proc = subprocess.Popen(gen_args)
    (out, err) = gen_proc.communicate()
	
    gen_args[1] = "Java\\build\\classes\\java\\main\\kernel.txt"
    gen_args[2] = sys.argv[4]
    gen_args[3] = sys.argv[4]	
    gen_proc = subprocess.Popen(gen_args)
	 
    (out, err) = gen_proc.communicate()
    args = ([
        "java",
        "-cp",
        "Java\\build\\classes\\java\\main\\",
		("Parallel" if len(sys.argv) == 7 else "") + "Main", 
		"Java\\build\\classes\\java\\main\data.txt", 
		"Java\\build\\classes\\java\\main\kernel.txt"
    ]) + ([] if len(sys.argv) < 7 else [sys.argv[6]]) + [
		"--log-to-socket"
	]
	
    print(args)
    avg_time = 0.0
    for i in range(int(sys.argv[5])):
        proc = subprocess.Popen(args)
		
		# remove comment if you want to manipulate program's output
        #(out, err) = proc.communicate()
        cl, new_addr = srv.accept()
        
		# we should receive a string formatted at "time 0.001"
        receipt = cl.recv(50)
        words = receipt.decode('utf-8').strip().split()
        cl.close()
        avg_time += float(words[1])

    avg_time /= int(sys.argv[5])

    log_to_excel(sys.argv[1], avg_time, '1' if len(sys.argv) < 7 else sys.argv[6], 'dynamic', sys.argv[2], sys.argv[3], sys.argv[4])