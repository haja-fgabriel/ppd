import os
import logging
import subprocess
import sys
from socket import *
import openpyxl

DEFAULT_IP = '127.0.0.1'
DEFAULT_PORT = 40000

def log_to_excel(filename, time, cmd_args):
    workbook = openpyxl.Workbook()
    workbook.active.append(('Run no.', 'Execution time', 'Command line args'))
    sheet1 = workbook.active
    
    last_run_no = sheet1.max_row
    sheet1.append((last_run_no, time, ' '.join(cmd_args)))
    sheet1['B' + str(sheet1.max_row)].number_format = '0.00000'
    workbook.save(filename)

if __name__ == '__main__':
    if len(sys.argv) < 4:
        exit(f"Usage: {sys.argv[0]} <excel_file> <number_runs> <process_name>")

    # special socket for getting logs from client
    srv = socket(AF_INET, SOCK_STREAM)
    srv.bind((DEFAULT_IP, DEFAULT_PORT))
    srv.listen(10)
    print(sys.argv[3:])

    avg_time = 0.0
    for i in range(int(sys.argv[2]) + 1):
        proc = subprocess.Popen(sys.argv[3:])
		
		# remove comment if you want to manipulate program's output
        #(out, err) = proc.communicate()
        cl, new_addr = srv.accept()
        
		# we should receive a string formatted at "time 0.001"
        receipt = cl.recv(50)
        words = receipt.decode('utf-8').strip().split()
        cl.close()
        avg_time += float(words[1])

    avg_time /= int(sys.argv[2])

    log_to_excel(sys.argv[1], words[1], sys.argv[3:])